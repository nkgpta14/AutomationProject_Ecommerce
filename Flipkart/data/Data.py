import openpyxl
from Flipkart.library import ConfigReader

path = ConfigReader.readConfigData('Details','File_location')
sheet_name = ConfigReader.readConfigData('Details','Sheet_Name')

def getRowCount():
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return (sheet.max_row)

def getColumnCount():
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return (sheet.max_column)

def readData(row_number,column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return (sheet.cell(row=row_number,column=column_number).value)

